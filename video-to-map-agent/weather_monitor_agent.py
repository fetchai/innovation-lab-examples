import os
import requests  # type: ignore[import-untyped]
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from uuid import uuid4
from dotenv import load_dotenv
from uagents import Agent, Context
from uagents_core.contrib.protocols.chat import (
    ChatMessage,
    TextContent,
    EndSessionContent,
)
from shared_models import (
    WeatherMonitorRequest,
    WeatherMonitorResponse,
    WeatherSnapshotRequest,
    WeatherSnapshotResponse,
)

# openpyxl is used only to append daily check rows to an existing workbook.
# Imported lazily so a missing install doesn't break the rest of the agent.
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except Exception:  # pragma: no cover
    load_workbook = None

load_dotenv()

agent = Agent(
    name="weather_monitor_agent",
    seed=os.getenv("WEATHER_SEED"),
    port=8008,
    endpoint=[os.getenv("WEATHER_MONITOR_ENDPOINT", "http://localhost:8008/submit")],
    network="testnet",
)

GOOGLE_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY")
WEATHER_BASE = "https://weather.googleapis.com/v1"


def get_daily_forecast(lat: float, lng: float, target_date: str) -> dict:
    try:
        url = (
            f"{WEATHER_BASE}/forecast/days:lookup"
            f"?key={GOOGLE_API_KEY}"
            f"&location.latitude={lat}"
            f"&location.longitude={lng}"
            f"&days=10"
        )
        data = requests.get(url, timeout=10).json()
        for day in data.get("forecastDays", []):
            display = day.get("displayDate", {})
            date_str = (
                f"{display.get('year')}-"
                f"{str(display.get('month')).zfill(2)}-"
                f"{str(display.get('day')).zfill(2)}"
            )
            if date_str == target_date:
                return day
    except Exception as e:
        print(f"Error fetching daily forecast: {e}")
    return {}


def get_public_alerts(lat: float, lng: float) -> list:
    try:
        url = (
            f"{WEATHER_BASE}/publicAlerts:lookup"
            f"?key={GOOGLE_API_KEY}"
            f"&location.latitude={lat}"
            f"&location.longitude={lng}"
        )
        data = requests.get(url, timeout=10).json()
        return data.get("alerts", [])
    except Exception:
        return []


def is_bad_forecast(day: dict) -> tuple:
    if not day:
        return False, ""
    daytime = day.get("daytimeForecast", {})
    condition_type = daytime.get("weatherCondition", {}).get("type", "")
    condition_desc = (
        daytime.get("weatherCondition", {}).get("description", {}).get("text", "")
    )
    precip_prob = (
        daytime.get("precipitation", {}).get("probability", {}).get("percent", 0) or 0
    )
    precip_mm = daytime.get("precipitation", {}).get("qpf", {}).get("quantity", 0) or 0
    thunder_prob = daytime.get("thunderstormProbability", 0) or 0
    wind_speed = daytime.get("wind", {}).get("speed", {}).get("value", 0) or 0

    bad_types = {
        "HEAVY_RAIN",
        "THUNDERSTORM",
        "TORNADO",
        "HURRICANE",
        "BLIZZARD",
        "HEAVY_SNOW",
        "TROPICAL_STORM",
        "FREEZING_RAIN",
        "ICE_STORM",
        "HAIL",
    }

    if condition_type in bad_types:
        return True, f"{condition_desc} forecast"
    if precip_prob >= 75 and precip_mm >= 15:
        return True, (
            f"{condition_desc} — {precip_prob}% chance, {precip_mm:.0f}mm expected"
        )
    if thunder_prob >= 60:
        return True, f"High thunderstorm probability ({thunder_prob}%)"
    if wind_speed >= 60:
        return True, f"High winds forecast ({wind_speed:.0f} km/h)"
    return False, ""


@agent.on_message(WeatherMonitorRequest)
async def handle_monitor_request(ctx: Context, sender: str, msg: WeatherMonitorRequest):
    ctx.logger.info(
        f"Weather monitoring started: {len(msg.stops)} stops, "
        f"trip date {msg.trip_start_date}"
    )
    stored = ctx.storage.get("trips") or {}
    stored[msg.user_sender_address] = {
        "stops": msg.stops,
        "trip_start_date": msg.trip_start_date,
        "user_sender_address": msg.user_sender_address,
        "alerted_stops": [],
        "excel_path": msg.excel_path,
    }
    ctx.storage.set("trips", stored)
    await ctx.send(sender, WeatherMonitorResponse(status="monitoring_started"))


# ── Excel log helpers ────────────────────────────────────────────────────────

_EXCEL_BORDER = None  # built lazily so openpyxl import failure is non-fatal


def _ensure_excel_style_cache():
    """openpyxl Border needs the module, and we don't want to import at
    top-level-failure cost. Build the reusable style objects once."""
    global _EXCEL_BORDER
    if _EXCEL_BORDER is None and load_workbook is not None:
        _EXCEL_BORDER = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )


def _forecast_summary_text(f: dict) -> str:
    if not f or not f.get("available"):
        return "Forecast unavailable"
    cond = f.get("condition") or ""
    parts = []
    if cond:
        parts.append(cond)
    high = f.get("high_c")
    low = f.get("low_c")
    if high is not None and low is not None:
        parts.append(f"{high:.0f}/{low:.0f} C")
    precip = f.get("precip_percent", 0) or 0
    if precip:
        parts.append(f"{precip}% rain")
    wind = f.get("wind_kmh", 0) or 0
    if wind >= 20:
        parts.append(f"wind {wind:.0f} km/h")
    return " | ".join(parts) if parts else "Details unavailable"


def _append_log_rows(excel_path: str, trip_date: str, per_stop_rows: list) -> None:
    """Append one row per stop to the 'Weather Log' sheet of the workbook.

    per_stop_rows: list of dicts with keys
        name, location, forecast, alert, recommendation
    """
    if not excel_path or load_workbook is None:
        return
    if not os.path.exists(excel_path):
        return

    _ensure_excel_style_cache()

    try:
        wb = load_workbook(excel_path)
        if "Weather Log" not in wb.sheetnames:
            return
        ws = wb["Weather Log"]

        # Find the next empty row by scanning column A starting below header.
        # Header is at row 4, data starts at row 5.
        start_row = 5
        row = start_row
        while ws.cell(row=row, column=1).value not in (None, ""):
            row += 1

        checked_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        for entry in per_stop_rows:
            values = [
                entry["name"],
                entry["location"],
                trip_date,
                checked_at,
                entry["forecast"],
                entry.get("alert", ""),
                entry["recommendation"],
            ]
            for col_idx, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=v)
                cell.font = Font(name="Calibri", size=10, color="193228")
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=(col_idx in (5, 6, 7)),
                )
                cell.border = _EXCEL_BORDER
            if entry.get("alert"):
                ws.cell(row=row, column=6).fill = PatternFill(
                    "solid",
                    fgColor="FEE2E2",
                )
                ws.cell(row=row, column=6).font = Font(
                    name="Calibri",
                    size=10,
                    bold=True,
                    color="B91C1C",
                )
            row += 1

        wb.save(excel_path)
    except Exception as e:
        # Never let excel bookkeeping break the actual monitoring.
        print(f"Error writing to excel: {e}")


@agent.on_interval(period=86400.0)
async def daily_weather_check(ctx: Context):
    ctx.logger.info("Running daily Google Weather API check")
    stored = ctx.storage.get("trips") or {}
    if not stored:
        ctx.logger.info("No trips to monitor")
        return

    updated = dict(stored)

    for user_addr, trip in stored.items():
        stops = trip["stops"]
        trip_date = trip["trip_start_date"]
        alerted = list(trip.get("alerted_stops", []))
        excel_path = trip.get("excel_path")
        new_alerts = []
        log_rows = []  # one per stop, for the Excel log sheet

        for i, stop in enumerate(stops):
            lat = stop["lat"]
            lng = stop["lng"]
            stop_name = stop["name"]
            stop_key = f"{i}_{stop_name}"

            # Always rebuild a forecast entry so the Excel log gets a
            # row for every stop every day, even if it's already been
            # alerted on (the log is a historical record).
            forecast_entry = _build_forecast_entry(stop, trip_date)
            summary = _forecast_summary_text(forecast_entry)
            alert_text = (
                forecast_entry.get("warning", "") if forecast_entry.get("bad") else ""
            )

            # Re-check public alerts — overrides forecast verdict.
            active_alerts = get_public_alerts(lat, lng)
            if active_alerts and not alert_text:
                alert_text = active_alerts[0].get(
                    "alertTitle",
                    "Active weather alert",
                )
                forecast_entry["bad"] = True

            recommendation = (
                "Consider rescheduling or an alternative"
                if forecast_entry.get("bad")
                else "OK to proceed"
            )
            if not forecast_entry.get("available"):
                recommendation = "Check back closer to the trip date"

            log_rows.append(
                {
                    "name": stop_name,
                    "location": f"{lat:.5f}, {lng:.5f}",
                    "forecast": summary,
                    "alert": alert_text,
                    "recommendation": recommendation,
                }
            )

            # Fire a chat alert only for newly-discovered bad stops.
            if forecast_entry.get("bad") and stop_key not in alerted:
                new_alerts.append(
                    {
                        "index": i + 1,
                        "name": stop_name,
                        "reason": alert_text or forecast_entry.get("warning", ""),
                    }
                )
                alerted.append(stop_key)

        # Write the per-stop log to the excel workbook regardless of
        # whether new alerts were generated.
        if excel_path:
            _append_log_rows(excel_path, trip_date, log_rows)

        if new_alerts:
            updated[user_addr]["alerted_stops"] = alerted

            lines = [f"Weather Alert — your trip on {trip_date}\n"]
            for alert in new_alerts:
                lines.append(
                    f"Stop {alert['index']} ({alert['name']}): "
                    f"{alert['reason']}. "
                    f"You may want to reschedule or find an alternative."
                )
            lines.append(
                "\nReply with 'suggest alternatives for stop [number]' "
                "and I'll find nearby options for you."
            )

            ctx.logger.info(f"Sending weather alert to {user_addr}")
            await ctx.send(
                user_addr,
                ChatMessage(
                    msg_id=uuid4(),
                    timestamp=datetime.now(timezone.utc),
                    content=[
                        TextContent(type="text", text="\n".join(lines)),
                        EndSessionContent(type="end-session"),
                    ],
                ),
            )

    ctx.storage.set("trips", updated)


# ── Initial snapshot (synchronous, one-shot) ──────────────────────────────────


def _build_forecast_entry(stop: dict, trip_date: str) -> dict:
    """Fetch the forecast for one stop and flatten the Google Weather
    response into a small, PDF/chat-friendly dict."""
    lat = stop["lat"]
    lng = stop["lng"]
    day = get_daily_forecast(lat, lng, trip_date)

    if not day:
        return {
            "name": stop.get("name", ""),
            "lat": lat,
            "lng": lng,
            "available": False,
            "condition": "",
            "high_c": None,
            "low_c": None,
            "precip_percent": 0,
            "wind_kmh": 0.0,
            "thunderstorm_percent": 0,
            "bad": False,
            "warning": "",
        }

    daytime = day.get("daytimeForecast", {}) or {}
    cond = daytime.get("weatherCondition", {}) or {}
    cond_text = (
        (cond.get("description", {}) or {}).get("text") or cond.get("type", "") or ""
    )

    def _num(obj, *keys):
        """Safely walk a nested dict and coerce the leaf to a number."""
        for k in keys:
            if not isinstance(obj, dict):
                return None
            obj = obj.get(k)
            if obj is None:
                return None
        try:
            return float(obj)
        except (TypeError, ValueError):
            return None

    high_c = _num(day, "maxTemperature", "degrees")
    low_c = _num(day, "minTemperature", "degrees")
    precip_percent = int(_num(daytime, "precipitation", "probability", "percent") or 0)
    wind_kmh = _num(daytime, "wind", "speed", "value") or 0.0
    thunder_percent = int(_num(daytime, "thunderstormProbability") or 0)

    bad, warning = is_bad_forecast(day)

    return {
        "name": stop.get("name", ""),
        "lat": lat,
        "lng": lng,
        "available": True,
        "condition": cond_text,
        "high_c": high_c,
        "low_c": low_c,
        "precip_percent": precip_percent,
        "wind_kmh": wind_kmh,
        "thunderstorm_percent": thunder_percent,
        "bad": bad,
        "warning": warning,
    }


@agent.on_message(WeatherSnapshotRequest)
async def handle_weather_snapshot(
    ctx: Context, sender: str, msg: WeatherSnapshotRequest
):
    ctx.logger.info(
        f"Weather snapshot requested for {len(msg.stops)} stops on "
        f"{msg.trip_start_date}"
    )
    if not msg.stops:
        await ctx.send(
            sender,
            WeatherSnapshotResponse(
                success=False,
                error="No stops provided.",
            ),
        )
        return

    try:
        # Google Weather calls are I/O-bound — fan out so a 10-stop snapshot
        # takes ~1s instead of ~10s.
        with ThreadPoolExecutor(max_workers=8) as pool:
            forecasts = list(
                pool.map(
                    lambda s: _build_forecast_entry(s, msg.trip_start_date),
                    msg.stops,
                )
            )
    except Exception as e:
        ctx.logger.error(f"Snapshot error: {e}")
        await ctx.send(
            sender,
            WeatherSnapshotResponse(
                success=False,
                error=str(e),
            ),
        )
        return

    unavailable = sum(1 for f in forecasts if not f["available"])
    bad = sum(1 for f in forecasts if f["bad"])
    ctx.logger.info(
        f"Snapshot ready: {len(forecasts)} stop(s), "
        f"{bad} flagged, {unavailable} unavailable."
    )
    await ctx.send(
        sender,
        WeatherSnapshotResponse(
            success=True,
            forecasts=forecasts,
        ),
    )


if __name__ == "__main__":
    agent.run()
