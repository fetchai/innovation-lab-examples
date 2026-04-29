"""
Excel Generator Agent — port 8012

Produces a multi-sheet .xlsx workbook alongside the PDF. Designed to be the
traveller's live day-to-day planning tool, while the PDF is the visual guide.

Sheets:
  1. Day-by-Day Itinerary      — sortable table of every stop per day
  2. Restaurant Guide          — every nearby restaurant found by Places
  3. Budget Tracker            — pre-filled estimates + editable Actual cols
  4. Weather Monitor Log       — initial forecast per stop, appended daily
                                 by the weather agent as it runs checks
  5. Maps & Links              — one row per stop with Google Maps URLs

Uses openpyxl (not the Cursor spreadsheets skill's JS artifact_tool) because
this file has to be generated dynamically at runtime by a running agent, not
once by an assistant.
"""

import os
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from uagents import Agent, Context
from shared_models import ExcelRequest, ExcelResponse

load_dotenv()
os.makedirs("output", exist_ok=True)

agent = Agent(
    name="excel_generator_agent",
    seed=os.getenv("EXCEL_SEED"),
    port=8012,
    endpoint=[os.getenv("EXCEL_GENERATOR_ENDPOINT", "http://localhost:8012/submit")],
    network="testnet",
)


# ── Colour palette (matches the PDF teal/coral theme) ────────────────────────

TEAL_HEX        = "1E9E75"
TEAL_LIGHT_HEX  = "D1F0E4"
CORAL_HEX       = "F5C4B3"
DARK_HEX        = "193228"
GREY_HEX        = "6B7280"

_HEADER_FILL = PatternFill("solid", fgColor=TEAL_HEX)
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_TITLE_FONT  = Font(name="Calibri", bold=True, size=16, color=DARK_HEX)
_SUBTITLE_FONT = Font(name="Calibri", italic=True, size=10, color=GREY_HEX)
_LABEL_FONT = Font(name="Calibri", bold=True, size=10, color=DARK_HEX)
_DAY_BAND_FILL = PatternFill("solid", fgColor="FFEFE6")
_TOTAL_FILL = PatternFill("solid", fgColor=TEAL_LIGHT_HEX)
_TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color=DARK_HEX)

_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _set_col_widths(ws, widths: dict) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _write_header_row(ws, row: int, headers: list) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws.row_dimensions[row].height = 22


def _format_data_cell(cell, *, bold=False, align="left", wrap=False):
    cell.font = Font(name="Calibri", size=10, bold=bold, color=DARK_HEX)
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap,
    )
    cell.border = _THIN_BORDER


def _add_table(ws, ref: str, name: str) -> None:
    """Wrap a header+data range in an Excel table for sort/filter in the UI."""
    try:
        tbl = Table(displayName=name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)
    except Exception:
        # Tables are a nice-to-have. If openpyxl chokes on a range edge case
        # we still have a perfectly usable grid.
        pass


def _flatten_planned_stops(planned_days: list) -> list:
    return [stop for day in planned_days for stop in day.get("stops", [])]


def _detect_dietary(preferences: str) -> str:
    if not preferences:
        return ""
    p = preferences.lower()
    for kw in ("vegan", "vegetarian", "halal", "kosher", "gluten-free", "pescatarian"):
        if kw in p:
            return kw
    return ""


def _forecast_summary(f: dict) -> str:
    """One-line forecast summary used in sheets 4 (weather log)."""
    if not f or not f.get("available"):
        return "Not yet available for this date"
    cond = f.get("condition") or ""
    parts = []
    if cond:
        parts.append(cond)
    high = f.get("high_c")
    low = f.get("low_c")
    if high is not None and low is not None:
        parts.append(f"{high:.0f}/{low:.0f} C")
    precip = f.get("precip_percent", 0) or 0
    if precip:
        parts.append(f"{precip}% rain")
    wind = f.get("wind_kmh", 0) or 0
    if wind >= 20:
        parts.append(f"wind {wind:.0f} km/h")
    return " | ".join(parts) if parts else "Details unavailable"


def _alert_text(f: dict) -> str:
    if not f or not f.get("available"):
        return ""
    return f.get("warning", "") if f.get("bad") else ""


def _recommendation(f: dict) -> str:
    if not f:
        return ""
    if not f.get("available"):
        return "Check back closer to the trip date"
    if f.get("bad"):
        return "Consider rescheduling or an alternative"
    return "OK to proceed"


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_cover_header(ws, msg: ExcelRequest) -> int:
    """Write title/meta block at the top of Sheet 1. Returns next free row."""
    ws.merge_cells("A1:H1")
    ws["A1"] = msg.trip_title or "Road Trip Itinerary"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    meta_bits = [f"Trip date: {msg.trip_start_date}"]
    meta_bits.append(
        f"{len(msg.planned_days)} day{'s' if len(msg.planned_days) != 1 else ''}"
    )
    meta_bits.append(f"{len(_flatten_planned_stops(msg.planned_days))} stops")
    meta_bits.append(f"Budget ~${msg.budget_per_day:.0f}/day")
    meta_bits.append(f"Total est. ${msg.total_estimated_cost:.0f}")

    ws.merge_cells("A2:H2")
    ws["A2"] = "  |  ".join(meta_bits)
    ws["A2"].font = _SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    if msg.video_title or msg.channel_name:
        source = f"Based on: {msg.video_title or ''}"
        if msg.channel_name:
            source += f" by {msg.channel_name}"
        ws.merge_cells("A3:H3")
        ws["A3"] = source
        ws["A3"].font = _SUBTITLE_FONT

    if msg.preferences:
        ws.merge_cells("A4:H4")
        ws["A4"] = f"Preferences: {msg.preferences}"
        ws["A4"].font = _SUBTITLE_FONT

    return 6  # leave a blank buffer row at 5


def _build_itinerary_sheet(ws, msg: ExcelRequest) -> None:
    header_row = _build_cover_header(ws, msg)

    headers = [
        "Day", "Theme", "Stop #", "Place Name", "Activity",
        "Duration (hrs)", "Est. Cost ($)", "GPS",
    ]
    _write_header_row(ws, header_row, headers)

    row = header_row + 1
    table_start = header_row
    for day in msg.planned_days:
        stops = day.get("stops", [])
        day_cost = float(day.get("estimated_cost_usd", 0) or 0)
        per_stop_cost = (day_cost / len(stops)) if stops else 0.0
        for stop_idx, stop in enumerate(stops, 1):
            values = [
                int(day.get("day_number", 0) or 0),
                day.get("theme", "") or "",
                stop_idx,
                stop.get("name", ""),
                stop.get("activity", "") or "",
                int(stop.get("duration_hours", 0) or 0),
                round(per_stop_cost, 2),
                f"{stop.get('lat', 0):.5f}, {stop.get('lng', 0):.5f}",
            ]
            for col_idx, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=v)
                _format_data_cell(cell, wrap=(col_idx in (2, 4, 5)))
            ws.cell(row=row, column=7).number_format = "$#,##0.00"
            row += 1

    table_end = row - 1
    if table_end > table_start:
        ref = f"A{table_start}:{get_column_letter(len(headers))}{table_end}"
        _add_table(ws, ref, "ItineraryTable")

    _set_col_widths(ws, {
        "A": 6, "B": 22, "C": 8, "D": 30, "E": 28,
        "F": 14, "G": 14, "H": 22,
    })
    ws.freeze_panes = f"A{header_row + 1}"


def _build_restaurants_sheet(ws, msg: ExcelRequest) -> None:
    dietary = _detect_dietary(msg.preferences or "") or "any"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Restaurant Guide"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Nearby restaurants returned by Google Places. "
        f"Dietary preference searched: {dietary}."
    )
    ws["A2"].font = _SUBTITLE_FONT

    headers = [
        "Stop", "Restaurant Name", "Rating", "Address",
        "Dietary Match", "Place Lat,Lng",
    ]
    header_row = 4
    _write_header_row(ws, header_row, headers)

    row = header_row + 1
    table_start = header_row
    any_rows = False
    for day in msg.planned_days:
        for stop in day.get("stops", []):
            stop_name = stop.get("name", "")
            stop_coords = f"{stop.get('lat', 0):.5f}, {stop.get('lng', 0):.5f}"
            restaurants = stop.get("nearby_restaurants") or []
            for r in restaurants:
                values = [
                    stop_name,
                    r.get("name", ""),
                    float(r.get("rating", 0) or 0),
                    r.get("address", ""),
                    dietary,
                    stop_coords,
                ]
                for col_idx, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=v)
                    _format_data_cell(cell, wrap=(col_idx in (2, 4)))
                ws.cell(row=row, column=3).number_format = "0.0"
                row += 1
                any_rows = True

    if not any_rows:
        ws.cell(row=row, column=1, value="(No nearby restaurants returned by Google Places)")
        ws.cell(row=row, column=1).font = Font(italic=True, color=GREY_HEX)
        row += 1

    table_end = row - 1
    if table_end > table_start and any_rows:
        ref = f"A{table_start}:{get_column_letter(len(headers))}{table_end}"
        _add_table(ws, ref, "RestaurantTable")

    _set_col_widths(ws, {
        "A": 24, "B": 32, "C": 8, "D": 40, "E": 16, "F": 22,
    })
    ws.freeze_panes = f"A{header_row + 1}"


def _build_budget_sheet(ws, msg: ExcelRequest) -> None:
    ws.merge_cells("A1:E1")
    ws["A1"] = "Budget Tracker"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        f"Per-day budget: ${msg.budget_per_day:.0f}. "
        "Update the 'Actual ($)' column as you spend."
    )
    ws["A2"].font = _SUBTITLE_FONT

    headers = ["Day", "Category", "Estimated ($)", "Actual ($)", "Notes"]
    header_row = 4
    _write_header_row(ws, header_row, headers)

    # Category split is a reasonable default; users can edit any cell.
    category_weights = [
        ("Food", 0.35),
        ("Entry fees / activities", 0.25),
        ("Gas / transport", 0.20),
        ("Lodging", 0.15),
        ("Misc", 0.05),
    ]

    row = header_row + 1
    data_start = row
    for day in msg.planned_days:
        day_num = int(day.get("day_number", 0) or 0)
        day_cost = float(day.get("estimated_cost_usd", 0) or 0)
        for category, weight in category_weights:
            ws.cell(row=row, column=1, value=day_num)
            ws.cell(row=row, column=2, value=category)
            est = round(day_cost * weight, 2)
            ws.cell(row=row, column=3, value=est)
            ws.cell(row=row, column=4, value=None)  # user fills in
            ws.cell(row=row, column=5, value=None)
            for c in range(1, 6):
                _format_data_cell(
                    ws.cell(row=row, column=c),
                    wrap=(c == 5),
                )
            ws.cell(row=row, column=3).number_format = "$#,##0.00"
            ws.cell(row=row, column=4).number_format = "$#,##0.00"
            row += 1

    data_end = row - 1

    # Totals row with live formulas.
    totals_row = row + 1
    ws.cell(row=totals_row, column=2, value="TOTAL").font = _TOTAL_FONT
    ws.cell(row=totals_row, column=2).alignment = Alignment(horizontal="right")
    if data_end >= data_start:
        ws.cell(
            row=totals_row, column=3,
            value=f"=SUM(C{data_start}:C{data_end})",
        )
        ws.cell(
            row=totals_row, column=4,
            value=f"=SUM(D{data_start}:D{data_end})",
        )
    else:
        ws.cell(row=totals_row, column=3, value=0)
        ws.cell(row=totals_row, column=4, value=0)
    for c in (2, 3, 4):
        ws.cell(row=totals_row, column=c).fill = _TOTAL_FILL
        ws.cell(row=totals_row, column=c).font = _TOTAL_FONT
        ws.cell(row=totals_row, column=c).border = _THIN_BORDER
    ws.cell(row=totals_row, column=3).number_format = "$#,##0.00"
    ws.cell(row=totals_row, column=4).number_format = "$#,##0.00"

    # Variance row.
    var_row = totals_row + 1
    ws.cell(row=var_row, column=2, value="Variance (Actual - Est.)")
    ws.cell(row=var_row, column=2).font = _LABEL_FONT
    ws.cell(row=var_row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(
        row=var_row, column=4,
        value=f"=D{totals_row}-C{totals_row}",
    )
    ws.cell(row=var_row, column=4).number_format = "$#,##0.00;[Red]-$#,##0.00"
    ws.cell(row=var_row, column=4).border = _THIN_BORDER

    _set_col_widths(ws, {
        "A": 6, "B": 30, "C": 16, "D": 16, "E": 32,
    })
    ws.freeze_panes = f"A{header_row + 1}"


def _build_weather_log_sheet(ws, msg: ExcelRequest) -> None:
    ws.merge_cells("A1:G1")
    ws["A1"] = "Weather Monitor Log"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G2")
    ws["A2"] = (
        "Initial forecast captured at generation time. The weather agent "
        "appends a new row per stop every 24h until the trip date."
    )
    ws["A2"].font = _SUBTITLE_FONT

    headers = [
        "Stop", "Location", "Trip Date", "Last Checked",
        "Forecast", "Alert", "Recommendation",
    ]
    header_row = 4
    _write_header_row(ws, header_row, headers)

    checked_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    row = header_row + 1
    table_start = header_row
    initial = msg.initial_forecasts or {}
    for day in msg.planned_days:
        for stop in day.get("stops", []):
            name = stop.get("name", "")
            loc = f"{stop.get('lat', 0):.5f}, {stop.get('lng', 0):.5f}"
            f = initial.get(name) or {}
            values = [
                name,
                loc,
                msg.trip_start_date,
                checked_at,
                _forecast_summary(f),
                _alert_text(f),
                _recommendation(f),
            ]
            for col_idx, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=v)
                _format_data_cell(cell, wrap=(col_idx in (5, 6, 7)))
            # Highlight alert cell red if present.
            if values[5]:
                ws.cell(row=row, column=6).fill = PatternFill(
                    "solid", fgColor="FEE2E2",
                )
                ws.cell(row=row, column=6).font = Font(
                    name="Calibri", size=10, bold=True, color="B91C1C",
                )
            row += 1

    table_end = row - 1
    if table_end > table_start:
        ref = f"A{table_start}:{get_column_letter(len(headers))}{table_end}"
        _add_table(ws, ref, "WeatherLogTable")

    _set_col_widths(ws, {
        "A": 26, "B": 22, "C": 14, "D": 18,
        "E": 36, "F": 30, "G": 34,
    })
    ws.freeze_panes = f"A{header_row + 1}"


def _build_maps_sheet(ws, msg: ExcelRequest) -> None:
    ws.merge_cells("A1:E1")
    ws["A1"] = "Maps & Links"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        "One Google Maps search link per stop (easy to open on mobile) "
        "plus the full road-trip route at the bottom."
    )
    ws["A2"].font = _SUBTITLE_FONT

    headers = ["Stop", "Name", "Address", "Google Maps Link", "Place ID"]
    header_row = 4
    _write_header_row(ws, header_row, headers)

    row = header_row + 1
    table_start = header_row
    stop_num = 1
    for day in msg.planned_days:
        for stop in day.get("stops", []):
            name = stop.get("name", "")
            place_id = stop.get("place_id", "")
            address = stop.get("address", "")
            # Build a per-stop search URL using Place ID when available.
            if place_id:
                link = (
                    "https://www.google.com/maps/search/?api=1"
                    f"&query={stop.get('lat', 0):.6f},{stop.get('lng', 0):.6f}"
                    f"&query_place_id={place_id}"
                )
            else:
                link = (
                    "https://www.google.com/maps/search/?api=1"
                    f"&query={stop.get('lat', 0):.6f},{stop.get('lng', 0):.6f}"
                )
            values = [stop_num, name, address, link, place_id]
            for col_idx, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=v)
                _format_data_cell(cell, wrap=(col_idx in (2, 3)))
            # Make the link a real hyperlink.
            link_cell = ws.cell(row=row, column=4)
            link_cell.hyperlink = link
            link_cell.font = Font(
                name="Calibri", size=10, color="1E40AF", underline="single",
            )
            stop_num += 1
            row += 1

    table_end = row - 1
    if table_end > table_start:
        ref = f"A{table_start}:{get_column_letter(len(headers))}{table_end}"
        _add_table(ws, ref, "MapsTable")

    # Full-route row below the table.
    row += 1
    ws.cell(row=row, column=1, value="FULL ROUTE").font = _TOTAL_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=2, value="Complete road trip (all stops in order)")
    ws.cell(row=row, column=2).font = _LABEL_FONT
    link_cell = ws.cell(row=row, column=4, value=msg.maps_url or "")
    if msg.maps_url:
        link_cell.hyperlink = msg.maps_url
    link_cell.font = Font(
        name="Calibri", size=10, bold=True, color="1E40AF", underline="single",
    )

    _set_col_widths(ws, {
        "A": 8, "B": 28, "C": 40, "D": 60, "E": 30,
    })
    ws.freeze_panes = f"A{header_row + 1}"


# ── Top-level builder ─────────────────────────────────────────────────────────

def generate_excel(msg: ExcelRequest) -> tuple[str, str]:
    wb = Workbook()

    # Sheet 1 starts as the default "Sheet".
    ws1 = wb.active
    ws1.title = "Itinerary"
    _build_itinerary_sheet(ws1, msg)

    _build_restaurants_sheet(wb.create_sheet("Restaurants"), msg)
    _build_budget_sheet(wb.create_sheet("Budget"), msg)
    _build_weather_log_sheet(wb.create_sheet("Weather Log"), msg)
    _build_maps_sheet(wb.create_sheet("Maps & Links"), msg)

    filename = f"itinerary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.abspath(os.path.join("output", filename))
    wb.save(filepath)
    return filename, filepath


@agent.on_message(ExcelRequest)
async def handle_excel(ctx: Context, sender: str, msg: ExcelRequest):
    total_stops = sum(len(d.get("stops", [])) for d in msg.planned_days)
    ctx.logger.info(
        f"Generating Excel workbook for {len(msg.planned_days)} day(s), "
        f"{total_stops} total stops"
    )
    try:
        filename, filepath = generate_excel(msg)
        ctx.logger.info(f"Excel saved: output/{filename}")
        await ctx.send(sender, ExcelResponse(
            success=True,
            excel_filename=filename,
            excel_path=filepath,
        ))
    except Exception as e:
        ctx.logger.error(f"Excel generation error: {e}")
        await ctx.send(sender, ExcelResponse(
            success=False,
            error=str(e),
        ))


if __name__ == "__main__":
    agent.run()
