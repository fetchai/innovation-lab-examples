"""
Multi-source parts sourcing service.

Sources tried in parallel (no proxy needed for most):
  1. PartSelect.com     — reliable HTML, good pricing
  2. AppliancePartsPros — good for appliance parts
  3. RepairClinic.com   — via Bright Data proxy if configured

All results are saved to an Excel file in reports/.
The best (cheapest) in-stock result is returned for the chat response.

Without Bright Data, direct HTTP requests are used (works for PartSelect + AppliancePartsPros).
With Bright Data, RepairClinic is also scraped.
"""

from __future__ import annotations

import asyncio
import json
import logging
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import httpx

from app.config.settings import get_settings

log = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
REPORTS_DIR = PROJECT_ROOT / "reports"

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────


async def fetch_parts_deterministic(
    part_name: str,
    part_number: str,
    context_text: str,
    brand: str = "",
    model_number: str = "",
    appliance_type: str = "",
) -> dict[str, float | str | list]:
    """
    Scrape multiple sources, save Excel report, return best result dict.

    Keys: price_usd, purchase_url, stock_status, source_site, all_sources (list), excel_path
    """
    results = await _scrape_all_sources(
        part_name,
        part_number,
        context_text,
        brand=brand,
        model_number=model_number,
        appliance_type=appliance_type,
    )
    excel_path = save_parts_excel(results, part_name, part_number) or ""

    # Pick cheapest in-stock result
    priced = [r for r in results if r["price_usd"] > 0]
    best = min(priced, key=lambda r: r["price_usd"]) if priced else results[0]

    return {
        "price_usd": best["price_usd"],
        "purchase_url": best["purchase_url"],
        "stock_status": best["stock_status"],
        "source_site": best["source_site"],
        "all_sources": results,
        "excel_path": excel_path,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Multi-source scraper
# ─────────────────────────────────────────────────────────────────────────────


def _smart_search_term(part_name: str, part_number: str, brand: str = "") -> str:
    """Build a search term that includes brand context for better relevance.
    Avoids duplicating the part number if it's already in part_name."""
    parts = [brand, part_name] if brand else [part_name]
    # Only append part_number if not already present in the string
    combined = " ".join(p for p in parts if p)
    if part_number.upper() not in combined.upper():
        combined = f"{combined} {part_number}"
    return combined.strip()


async def _scrape_all_sources(
    part_name: str,
    part_number: str,
    context_text: str,
    brand: str = "",
    model_number: str = "",
    appliance_type: str = "",
) -> list[dict]:
    """Run all scrapers concurrently and return a merged ranked list."""
    settings = get_settings()
    proxy = settings.brightdata_proxy_url

    # Build a richer search term that includes brand for sites where it helps
    search_hint = _smart_search_term(part_name, part_number, brand)

    tasks = [
        _scrape_appliancepartspros(part_number, part_name),
        _scrape_ebay(part_number, search_hint),
        _scrape_applianceparts365(part_number, part_name),
        _scrape_partselect(part_number, part_name, proxy),
        _scrape_repairclinic(part_number, part_name, proxy),
        _scrape_amazon(part_number, search_hint, proxy),
    ]

    raw = await asyncio.gather(*tasks, return_exceptions=True)

    results: list[dict] = []
    for r in raw:
        if isinstance(r, Exception):
            log.debug("Source scrape error: %s", r)
            continue
        if isinstance(r, list):
            results.extend(r)
        elif isinstance(r, dict):
            results.append(r)

    if not results:
        results = [_make_stub(part_number, part_name)]

    # Deduplicate: same store + same price is almost certainly the same listing
    seen: set[tuple[str, float]] = set()
    deduped: list[dict[str, object]] = []
    for entry in results:
        key = (str(entry["source_site"]), round(float(entry["price_usd"]), 2))  # type: ignore[arg-type]
        if key not in seen:
            seen.add(key)
            deduped.append(entry)
    results = deduped  # type: ignore[assignment]

    # Sort: priced in-stock first, then others
    results.sort(key=lambda x: (float(x["price_usd"]) <= 0, float(x["price_usd"])))  # type: ignore[arg-type]
    return results[:10]  # cap at 10


# ─────────────────────────────────────────────────────────────────────────────
# Individual site scrapers
# ─────────────────────────────────────────────────────────────────────────────


def _page_has_part(html: str, part_number: str) -> bool:
    """
    Returns True if the scraped page actually contains the searched part number.
    This prevents returning prices for unrelated parts that happen to appear on
    category/catalog browse pages (e.g. RepairClinic's 298k-result pages).
    """
    if not part_number or not part_number.strip():
        return True
    return part_number.upper() in html.upper()


def _extract_json_ld_prices(
    html: str, site: str, fallback_url: str, part_number: str = ""
) -> list[dict]:
    """
    Extract product prices from JSON-LD structured data blocks.

    When part_number is provided, only accepts products whose identifiers
    (name, sku, mpn, productID) contain the part number — preventing false
    matches from category pages or multi-product JSON-LD blocks.
    """
    results = []
    pn_upper = part_number.upper() if part_number else ""

    for m in re.finditer(
        r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        html,
        re.DOTALL | re.IGNORECASE,
    ):
        try:
            data = json.loads(m.group(1))
        except (json.JSONDecodeError, ValueError):
            continue

        raw = data if isinstance(data, list) else [data]
        # Flatten @graph nodes so they are also visited
        items: list = []
        for node in raw:
            if isinstance(node, dict) and "@graph" in node:
                items.extend(node["@graph"])
            else:
                items.append(node)
        for obj in items:
            if not isinstance(obj, dict):
                continue
            t = obj.get("@type", "")
            if t not in ("Product", "ItemPage", "Offer") and "Product" not in str(t):
                continue

            # ── Validate this JSON-LD product is for our specific part ──────
            if pn_upper:
                product_identity = " ".join(
                    filter(
                        None,
                        [
                            str(obj.get("name", "")),
                            str(obj.get("sku", "")),
                            str(obj.get("mpn", "")),
                            str(obj.get("productID", "")),
                            str(obj.get("identifier", "")),
                            str(obj.get("gtin", "")),
                        ],
                    )
                ).upper()
                if pn_upper not in product_identity:
                    log.debug(
                        "[json-ld] Skipping product '%s' — part number %s not found in identifiers",
                        obj.get("name", "?")[:60],
                        part_number,
                    )
                    continue

            offers = obj.get("offers") or obj.get("Offers") or {}
            if isinstance(offers, dict):
                offers = [offers]
            elif not isinstance(offers, list):
                offers = [obj]
            for offer in offers:
                try:
                    price = float(str(offer.get("price") or offer.get("lowPrice") or 0))
                    if price < 1 or price > 3000:
                        continue
                    p_url = offer.get("url") or obj.get("url") or fallback_url
                    avail = str(offer.get("availability", "")).lower()
                    stock = (
                        "In Stock"
                        if "instock" in avail
                        else "Out of Stock"
                        if "outofstock" in avail
                        else "Check Vendor"
                    )
                    results.append(
                        {
                            "source_site": site,
                            "price_usd": price,
                            "purchase_url": p_url,
                            "stock_status": stock,
                        }
                    )
                except (TypeError, ValueError):
                    continue

    return results


async def _scrape_partselect(
    part_number: str, part_name: str, proxy: str | None
) -> list[dict]:
    """
    PartSelect.com — routes through Bright Data proxy when available.

    PartSelect returns 403 to direct bot requests. Even with proxy, if the search
    results redirect to a category page (not a specific part page), the part number
    won't appear on the page → we return a search link only (no false price).
    """
    url = f"https://www.partselect.com/Search.aspx?SearchTerm={quote(part_number)}"
    try:
        if proxy:
            transport = httpx.AsyncHTTPTransport(
                proxy=httpx.Proxy(url=proxy), verify=False
            )
            client = httpx.AsyncClient(
                transport=transport,
                headers=_HEADERS,
                timeout=45.0,
                follow_redirects=True,
            )
        else:
            client = httpx.AsyncClient(
                headers=_HEADERS, timeout=20.0, follow_redirects=True
            )
        async with client as c:
            r = await c.get(url)
            r.raise_for_status()
            html = r.text
    except Exception as e:
        msg = str(e)
        if "bad_endpoint" in msg or "robots.txt" in msg:
            log.warning(
                "[partselect] Bright Data zone does not have PartSelect access "
                "(bad_endpoint/robots.txt). Ask your account manager to enable it. "
                "Returning search link."
            )
        elif "407" in msg or "Proxy" in msg.lower():
            log.warning(
                "[partselect] Bright Data proxy auth error — check BRIGHTDATA credentials."
            )
        else:
            log.debug("PartSelect failed: %s", e)
        return [_link_stub(part_number, part_name, "partselect.com", url)]

    # Gate: only extract prices if the page is actually about our part
    if not _page_has_part(html, part_number):
        log.debug(
            "[partselect] Part %s not on page — returning search link", part_number
        )
        return [_link_stub(part_number, part_name, "partselect.com", url)]

    # 1 — JSON-LD with part-number validation
    results = _extract_json_ld_prices(html, "partselect.com", url, part_number)

    # 2 — itemprop price near our part number
    if not results:
        # Find all occurrences of the part number on the page, look for nearby price
        for pn_m in re.finditer(re.escape(part_number), html, re.IGNORECASE):
            window = html[max(0, pn_m.start() - 200) : pn_m.start() + 600]
            price_m = re.search(
                r'itemprop=["\']price["\'][^>]*content=["\']([0-9]+\.?[0-9]*)["\']'
                r'|(?:class="[^"]*price[^"]*"|>\$)([0-9]+\.[0-9]{2})',
                window,
                re.IGNORECASE,
            )
            if price_m:
                try:
                    price = float(price_m.group(1) or price_m.group(2))
                    if 1 < price < 3000:
                        url_m = re.search(
                            r'href="(/[A-Za-z0-9\-]+-PS[0-9]+[^"]*)"', window
                        )
                        part_url = (
                            ("https://www.partselect.com" + url_m.group(1))
                            if url_m
                            else url
                        )
                        stock = "In Stock" if "In Stock" in window else "Check Vendor"
                        results.append(
                            {
                                "source_site": "partselect.com",
                                "price_usd": price,
                                "purchase_url": part_url,
                                "stock_status": stock,
                            }
                        )
                        break
                except (TypeError, ValueError):
                    continue

    if not results:
        results = [_link_stub(part_number, part_name, "partselect.com", url)]
    return results[:3]


async def _scrape_appliancepartspros(part_number: str, part_name: str) -> list[dict]:
    """AppliancePartsPros — JSON-LD + dollar-regex fallback with part-number gate."""
    url = f"https://www.appliancepartspros.com/search.aspx?q={quote(part_number)}"
    try:
        async with httpx.AsyncClient(
            headers=_HEADERS, timeout=20.0, follow_redirects=True
        ) as c:
            r = await c.get(url)
            r.raise_for_status()
            html = r.text
    except Exception as e:
        log.debug("AppliancePartsPros failed: %s", e)
        return [_link_stub(part_number, part_name, "appliancepartspros.com", url)]

    # Gate: only trust prices if the page mentions our part number
    if not _page_has_part(html, part_number):
        log.debug("[APP] Part %s not on page — returning search link", part_number)
        return [_link_stub(part_number, part_name, "appliancepartspros.com", url)]

    # 1 — JSON-LD with part-number validation
    results = _extract_json_ld_prices(html, "appliancepartspros.com", url, part_number)

    # 2 — dollar-regex near the part number occurrence
    if not results:
        for pn_m in re.finditer(re.escape(part_number), html, re.IGNORECASE):
            window = html[max(0, pn_m.start() - 100) : pn_m.start() + 800]
            price_m = re.search(r"\$([0-9]+\.[0-9]{2})", window)
            if price_m:
                try:
                    price = float(price_m.group(1))
                    if 5 < price < 2000:
                        stock = (
                            "In Stock" if "Add to Cart" in window else "Check Vendor"
                        )
                        results.append(
                            {
                                "source_site": "appliancepartspros.com",
                                "price_usd": price,
                                "purchase_url": url,
                                "stock_status": stock,
                            }
                        )
                        break
                except (TypeError, ValueError):
                    continue

    # 3 — first dollar amount on page (last resort — APP usually redirects to exact part)
    if not results:
        for m in re.finditer(r"\$([0-9]+\.[0-9]{2})", html):
            try:
                price = float(m.group(1))
                if price < 5 or price > 2000:
                    continue
                stock = (
                    "In Stock"
                    if "Add to Cart" in html[max(0, m.start() - 300) : m.start() + 300]
                    else "Check Vendor"
                )
                results.append(
                    {
                        "source_site": "appliancepartspros.com",
                        "price_usd": price,
                        "purchase_url": url,
                        "stock_status": stock,
                    }
                )
                break
            except (TypeError, ValueError):
                continue

    if not results:
        results = [_link_stub(part_number, part_name, "appliancepartspros.com", url)]
    return results[:3]


async def _scrape_ebay(part_number: str, part_name: str) -> list[dict]:
    """
    eBay — multiple extraction strategies, all anchored to the part number.

    eBay search results embed the part number in listing titles when the seller
    includes it, so the page-level gate is reliable here.
    """
    # Build search: use part_name directly if it already contains the part
    # number (e.g., search_hint = "GE Evaporator Fan Motor WR60X26866").
    if part_number.upper() in part_name.upper():
        search_term = part_name
    else:
        search_term = f"{part_number} {part_name}".strip()
    url = f"https://www.ebay.com/sch/i.html?_nkw={quote(search_term)}&LH_BIN=1&_sop=15"
    try:
        async with httpx.AsyncClient(
            headers=_HEADERS, timeout=20.0, follow_redirects=True
        ) as c:
            r = await c.get(url)
            r.raise_for_status()
            html = r.text
            final_url = str(r.url)
    except Exception as e:
        log.debug("eBay failed: %s", e)
        return [_link_stub(part_number, part_name, "ebay.com", url)]

    # Gate: detect eBay bot challenge / CAPTCHA redirects.
    # eBay redirects bots to splashui/challenge — the HTML contains no
    # real listings, just random page furniture with dollar signs that
    # the regex would pick up as fake prices.
    if "splashui/challenge" in final_url or "captcha" in html.lower():
        log.warning(
            "[eBay] Bot challenge detected (redirected to %s) — "
            "returning search link instead of fake prices.",
            final_url[:80],
        )
        return [_link_stub(part_number, part_name, "ebay.com", url)]

    if not _page_has_part(html, part_number):
        log.debug(
            "[eBay] Part %s not in search results — returning search link", part_number
        )
        return [_link_stub(part_number, part_name, "ebay.com", url)]

    results = []

    # 1 — JSON-LD with part-number validation
    results = _extract_json_ld_prices(html, "ebay.com", url, part_number)

    # 2 — Find price near each occurrence of the part number in the listing block
    if not results:
        for pn_m in re.finditer(re.escape(part_number), html, re.IGNORECASE):
            # Expand context: eBay item blocks can be 2-3kb
            window = html[max(0, pn_m.start() - 1500) : pn_m.start() + 500]
            # Look for price span
            price_m = re.search(
                r'class="s-item__price"[^>]*>\s*(?:US )?\$([0-9]+(?:\.[0-9]+)?)'
                r'|"US \$([0-9]+\.[0-9]{2})"'
                r"|(?:US )?\$([0-9]+\.[0-9]{2})",
                window,
            )
            if price_m:
                try:
                    price = float(
                        price_m.group(1) or price_m.group(2) or price_m.group(3)
                    )
                    if 5 < price < 2000:
                        url_m = re.search(
                            r'href="(https://www\.ebay\.com/itm/[^"?]+)', window
                        )
                        item_url = url_m.group(1) if url_m else url
                        results.append(
                            {
                                "source_site": "ebay.com",
                                "price_usd": price,
                                "purchase_url": item_url,
                                "stock_status": "Buy It Now",
                            }
                        )
                        if len(results) >= 3:
                            break
                except (TypeError, ValueError):
                    continue

    # 3 — s-item__price spans (picks up correctly ordered results)
    if not results:
        for m in re.finditer(
            r'class="s-item__price">\s*(?:US )?\$([0-9]+(?:\.[0-9]+)?)', html
        ):
            try:
                price = float(m.group(1))
                if price < 5 or price > 2000:
                    continue
                nearby = html[max(0, m.start() - 1500) : m.start()]
                # Only accept if the part number appeared in this listing block
                if part_number.upper() not in nearby.upper():
                    continue
                url_m = re.search(r'href="(https://www\.ebay\.com/itm/[^"?]+)', nearby)
                item_url = url_m.group(1) if url_m else url
                results.append(
                    {
                        "source_site": "ebay.com",
                        "price_usd": price,
                        "purchase_url": item_url,
                        "stock_status": "Buy It Now",
                    }
                )
                if len(results) >= 3:
                    break
            except (TypeError, ValueError):
                continue

    if not results:
        results = [_link_stub(part_number, part_name, "ebay.com", url)]
    return results[:3]


async def _scrape_applianceparts365(part_number: str, part_name: str) -> list[dict]:
    """
    ApplianceParts365.com — nopCommerce store with a direct search endpoint.
    URL: /search?q={part_number}

    ApplianceParts365 shows sale prices like: ~~$100.38~~ $62.66
    The old price is in a <del>/<s>/strikethrough element, the sale price follows.
    We must pick the ACTUAL price (lowest displayed), not the strikethrough.
    """
    url = f"https://applianceparts365.com/search?q={quote(part_number)}"
    try:
        async with httpx.AsyncClient(
            headers=_HEADERS, timeout=20.0, follow_redirects=True
        ) as c:
            r = await c.get(url)
            r.raise_for_status()
            html = r.text
    except Exception as e:
        log.debug("ApplianceParts365 failed: %s", e)
        return [_link_stub(part_number, part_name, "applianceparts365.com", url)]

    if not _page_has_part(html, part_number):
        log.debug("[AP365] Part %s not on page — returning search link", part_number)
        return [_link_stub(part_number, part_name, "applianceparts365.com", url)]

    # 1 — JSON-LD with part-number validation
    results = _extract_json_ld_prices(html, "applianceparts365.com", url, part_number)

    # 2 — price near part number, preferring sale/actual price over strikethrough
    if not results:
        for pn_m in re.finditer(re.escape(part_number), html, re.IGNORECASE):
            window = html[max(0, pn_m.start() - 100) : pn_m.start() + 800]

            # Find ALL dollar amounts in the window
            all_prices = [
                float(m.group(1))
                for m in re.finditer(r"\$([0-9]+\.[0-9]{2})", window)
                if 5 < float(m.group(1)) < 2000
            ]
            if not all_prices:
                continue

            # The actual/sale price is typically the LOWEST displayed price.
            # The strikethrough price is the higher original price.
            price = min(all_prices)

            stock = (
                "In Stock"
                if "Add to Cart" in window or "Buy" in window
                else "Check Vendor"
            )
            url_m = re.search(
                r'href="(/[^"]+' + re.escape(part_number.lower()) + r'[^"]*)"',
                window,
                re.IGNORECASE,
            )
            part_url = (
                ("https://applianceparts365.com" + url_m.group(1)) if url_m else url
            )
            results.append(
                {
                    "source_site": "applianceparts365.com",
                    "price_usd": price,
                    "purchase_url": part_url,
                    "stock_status": stock,
                }
            )
            break

    if not results:
        results = [_link_stub(part_number, part_name, "applianceparts365.com", url)]
    return results[:3]


async def _scrape_amazon(
    part_number: str, part_name: str, proxy: str | None
) -> list[dict]:
    """
    Amazon.com — requires Bright Data proxy (blocks all direct bot traffic).

    Amazon search URL: /s?k={part_number}
    Price extraction uses span.a-offscreen (the screen-reader price, always in $XX.XX format)
    and JSON-LD structured data embedded on search result pages.
    """
    # Build search term: avoid duplicating part_number if it's already in part_name
    amz_search = (
        f"{part_number} {part_name}"
        if part_number.upper() not in part_name.upper()
        else part_name
    )
    url = f"https://www.amazon.com/s?k={quote(amz_search.strip())}"

    if not proxy:
        # Direct access always fails — just give a valid search link
        return [
            {
                "source_site": "amazon.com",
                "price_usd": 0.0,
                "purchase_url": url,
                "stock_status": "Click to check price",
            }
        ]

    log.info("Bright Data → Amazon: %s", url)
    try:
        transport = httpx.AsyncHTTPTransport(proxy=httpx.Proxy(url=proxy), verify=False)
        async with httpx.AsyncClient(
            transport=transport, headers=_HEADERS, timeout=60.0, follow_redirects=True
        ) as c:
            r = await c.get(url)
            r.raise_for_status()
            html = r.text
    except Exception as e:
        log.warning("Amazon via Bright Data failed: %s — returning search link", e)
        return [
            {
                "source_site": "amazon.com",
                "price_usd": 0.0,
                "purchase_url": url,
                "stock_status": "Click to check price",
            }
        ]

    if not _page_has_part(html, part_number):
        log.debug("[Amazon] Part %s not in search results", part_number)
        return [
            {
                "source_site": "amazon.com",
                "price_usd": 0.0,
                "purchase_url": url,
                "stock_status": "Click to check price",
            }
        ]

    results = []

    # 1 — JSON-LD with part-number validation
    results = _extract_json_ld_prices(html, "amazon.com", url, part_number)

    # 2 — a-offscreen span near the part number (Amazon's machine-readable price)
    # Format: <span class="a-offscreen">$XX.XX</span> inside a-price container
    if not results:
        for pn_m in re.finditer(re.escape(part_number), html, re.IGNORECASE):
            window = html[max(0, pn_m.start() - 2000) : pn_m.start() + 2000]
            # a-offscreen is the cleanest price signal on Amazon
            price_m = re.search(
                r'class="a-offscreen">\s*\$([0-9]+(?:\.[0-9]+)?)', window
            )
            if price_m:
                try:
                    price = float(price_m.group(1))
                    if 5 < price < 2000:
                        # Find the ASIN-based product URL
                        asin_m = re.search(r"/dp/([A-Z0-9]{10})", window)
                        item_url = (
                            f"https://www.amazon.com/dp/{asin_m.group(1)}"
                            if asin_m
                            else url
                        )
                        results.append(
                            {
                                "source_site": "amazon.com",
                                "price_usd": price,
                                "purchase_url": item_url,
                                "stock_status": "In Stock",
                            }
                        )
                        if len(results) >= 3:
                            break
                except (TypeError, ValueError):
                    continue

    # 3 — Any a-offscreen on the page (broader fallback)
    if not results:
        for m in re.finditer(r'class="a-offscreen">\s*\$([0-9]+\.[0-9]{2})', html):
            try:
                price = float(m.group(1))
                if 5 < price < 2000:
                    nearby = html[max(0, m.start() - 2000) : m.start() + 500]
                    if part_number.upper() not in nearby.upper():
                        continue  # Only accept prices in blocks containing our part number
                    asin_m = re.search(r"/dp/([A-Z0-9]{10})", nearby)
                    item_url = (
                        f"https://www.amazon.com/dp/{asin_m.group(1)}"
                        if asin_m
                        else url
                    )
                    results.append(
                        {
                            "source_site": "amazon.com",
                            "price_usd": price,
                            "purchase_url": item_url,
                            "stock_status": "In Stock",
                        }
                    )
                    if len(results) >= 3:
                        break
            except (TypeError, ValueError):
                continue

    if not results:
        results = [
            {
                "source_site": "amazon.com",
                "price_usd": 0.0,
                "purchase_url": url,
                "stock_status": "Click to check price",
            }
        ]
    return results[:3]


async def _scrape_repairclinic(
    part_number: str, part_name: str, proxy: str | None
) -> list[dict]:
    """
    RepairClinic.com scraper.

    RepairClinic is an Angular SPA. The correct server-side search URL is:
      /Shop-For-Parts?SearchTerm=<part>
    With Bright Data Web Unlocker the page renders fully and JSON-LD is present.
    Without proxy we still provide a valid clickable search link.
    """
    search_url = (
        f"https://www.repairclinic.com/Shop-For-Parts?SearchTerm={quote(part_number)}"
    )

    if not proxy:
        return [
            {
                "source_site": "repairclinic.com",
                "price_usd": 0.0,
                "purchase_url": search_url,
                "stock_status": "Click to check price",
            }
        ]

    log.info("Bright Data → RepairClinic: %s", search_url)
    try:
        transport = httpx.AsyncHTTPTransport(proxy=httpx.Proxy(url=proxy), verify=False)
        async with httpx.AsyncClient(
            transport=transport, headers=_HEADERS, timeout=60.0, follow_redirects=True
        ) as c:
            r = await c.get(search_url)
            r.raise_for_status()
            html = r.text
    except Exception as e:
        log.warning(
            "RepairClinic via Bright Data failed: %s — returning search link", e
        )
        return [
            {
                "source_site": "repairclinic.com",
                "price_usd": 0.0,
                "purchase_url": search_url,
                "stock_status": "Visit site for pricing",
            }
        ]

    # Gate 1: part number must appear on the page (not just in the search bar)
    if not _page_has_part(html, part_number):
        log.warning(
            "[RepairClinic] Part %s not found in rendered HTML — "
            "page may be a generic catalog browse. Returning search link.",
            part_number,
        )
        return [
            {
                "source_site": "repairclinic.com",
                "price_usd": 0.0,
                "purchase_url": search_url,
                "stock_status": "Visit site for pricing",
            }
        ]

    # Gate 2: detect garbage catalog pages — if result count is huge, the search
    # didn't match a specific part (e.g., "261,597 results" = generic browse)
    result_count_m = re.search(r"([\d,]+)\s+results?", html)
    if result_count_m:
        count = int(result_count_m.group(1).replace(",", ""))
        if count > 100:
            log.warning(
                "[RepairClinic] Search returned %d results for %s — too broad, "
                "not a specific part match. Returning search link.",
                count,
                part_number,
            )
            return [
                {
                    "source_site": "repairclinic.com",
                    "price_usd": 0.0,
                    "purchase_url": search_url,
                    "stock_status": "Visit site for pricing",
                }
            ]

    # 1 — JSON-LD with part-number validation (prevents rack adjuster / wrong-part matches)
    results = _extract_json_ld_prices(html, "repairclinic.com", search_url, part_number)

    # 2 — data-price attribute near part number
    if not results:
        for pn_m in re.finditer(re.escape(part_number), html, re.IGNORECASE):
            window = html[max(0, pn_m.start() - 300) : pn_m.start() + 800]
            dp_m = re.search(
                r'data-price="([0-9]+(?:\.[0-9]+)?)"[^>]*href="([^"]+repairclinic[^"]+)"',
                window,
            )
            if dp_m:
                try:
                    results.append(
                        {
                            "source_site": "repairclinic.com",
                            "price_usd": float(dp_m.group(1)),
                            "purchase_url": dp_m.group(2),
                            "stock_status": "In Stock"
                            if "Add to Cart" in window
                            else "Check Vendor",
                        }
                    )
                    break
                except ValueError:
                    continue

    # 3 — dollar amount near part number
    if not results:
        for pn_m in re.finditer(re.escape(part_number), html, re.IGNORECASE):
            window = html[max(0, pn_m.start() - 100) : pn_m.start() + 600]
            price_m = re.search(r"\$([0-9]+\.[0-9]{2})", window)
            if price_m:
                try:
                    price = float(price_m.group(1))
                    if 5 < price < 2000:
                        results.append(
                            {
                                "source_site": "repairclinic.com",
                                "price_usd": price,
                                "purchase_url": search_url,
                                "stock_status": "Check Vendor",
                            }
                        )
                        break
                except (TypeError, ValueError):
                    continue

    if not results:
        results = [
            {
                "source_site": "repairclinic.com",
                "price_usd": 0.0,
                "purchase_url": search_url,
                "stock_status": "Visit site for pricing",
            }
        ]
    return results[:3]


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────


def save_parts_excel(
    results: list[dict], part_name: str, part_number: str
) -> str | None:
    """Save results to an Excel file in reports/. Returns the file path."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning(
            "openpyxl not installed — skipping Excel export. Run: pip install openpyxl"
        )
        return None

    REPORTS_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_pn = re.sub(r"[^\w]", "_", part_number)
    path = REPORTS_DIR / f"parts_{safe_pn}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parts Price Comparison"

    # Header row
    headers = ["Rank", "Source", "Price (USD)", "Stock Status", "Buy URL"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Title row above headers
    ws.insert_rows(1)
    title_cell = ws.cell(
        row=1, column=1, value=f"Price Comparison: {part_name} ({part_number})"
    )
    title_cell.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    best_fill = PatternFill("solid", fgColor="E2EFDA")
    other_fill = PatternFill("solid", fgColor="FFFFFF")

    for i, src in enumerate(results, 1):
        row = i + 2
        fill = best_fill if i == 1 else other_fill
        rank_label = f"#{i} ★ BEST" if i == 1 else f"#{i}"
        ws.cell(row=row, column=1, value=rank_label).fill = fill
        ws.cell(row=row, column=2, value=src.get("source_site", "")).fill = fill
        price_cell = ws.cell(row=row, column=3, value=src.get("price_usd", 0))
        price_cell.number_format = '"$"#,##0.00'
        price_cell.fill = fill
        ws.cell(row=row, column=4, value=src.get("stock_status", "")).fill = fill
        url_val = src.get("purchase_url", "")
        url_cell = ws.cell(row=row, column=5, value=url_val)
        url_cell.fill = fill
        if url_val and url_val.startswith("http"):
            url_cell.hyperlink = url_val
            url_cell.font = Font(color="0563C1", underline="single")

    # Column widths
    widths = [12, 22, 14, 20, 60]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(path)
    log.info("Excel saved: %s", path)
    return str(path)


# ─────────────────────────────────────────────────────────────────────────────
# Stub helpers
# ─────────────────────────────────────────────────────────────────────────────


def _make_stub(part_number: str, part_name: str) -> dict:
    return _link_stub(
        part_number,
        part_name,
        "partselect.com",
        f"https://www.partselect.com/Search.aspx?SearchTerm={quote(part_number)}",
    )


def _link_stub(part_number: str, part_name: str, site: str, url: str) -> dict:
    return {
        "source_site": site,
        "price_usd": 0.0,
        "purchase_url": url,
        "stock_status": "Click to check price",
    }


# Back-compat shim
async def fetch_part_price_and_url(
    part_number: str, appliance_hint: str = ""
) -> tuple[float, str, str]:
    d = await fetch_parts_deterministic(part_number, part_number, appliance_hint)
    return float(d["price_usd"]), str(d["purchase_url"]), str(d["stock_status"])  # type: ignore[arg-type]  # type: ignore[arg-type]
